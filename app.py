from flask import Flask, request, jsonify, render_template, send_from_directory
import os
import openpyxl
import time
import imagehash
from PIL import Image
from tqdm import tqdm

app = Flask(__name__, static_url_path='/static')  # Serve static files from 'static' directory
EXCEL_FILE = "file_list.xlsx" # Define Excel file name
ROOT_DIRECTORY = ""

def list_files_and_find_duplicates(root_dir, excel_file, similarity_threshold=5):
    """
    Lists files, identifies duplicates, groups them, and creates/updates the Excel file.
    """

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    header = ["File Name", "Directory Path", "File Path", "File Size (bytes)", "Last Modified", "Duplicate Group", "Delete"]
    sheet.append(header)

    file_data = []  # Store file data before writing
    image_hashes = {}
    duplicate_groups = {}

    total_files = 0
    for dirpath, dirnames, filenames in os.walk(root_dir):
        total_files += len(filenames)

    with tqdm(total=total_files, desc="Listing and Analyzing Files", unit="file", disable=True) as pbar: #disable tqdm on web
        for dirpath, dirnames, filenames in os.walk(root_dir):
            for filename in filenames:
                filepath = os.path.join(dirpath, filename)
                if is_image_or_video(filename):
                    try:
                        file_size = os.path.getsize(filepath)
                        modified_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(os.path.getmtime(filepath)))

                        try:
                            img = Image.open(filepath)
                            hash_value = imagehash.average_hash(img)

                            # Check for duplicates
                            found_duplicate = False
                            for existing_hash, file_list in duplicate_groups.items():
                                if hash_value - existing_hash < similarity_threshold:
                                    duplicate_groups[existing_hash].append(filepath)
                                    found_duplicate = True
                                    break

                            if not found_duplicate:
                                duplicate_groups[hash_value] = [filepath]

                        except Exception as e:
                            print(f"Warning: Error processing image {filepath} for duplicate detection: {e}")
                            hash_value = None # Indicate error during hash calculation

                        directory_path, file_name = os.path.split(filepath) # split the paths
                        # Shorten the directory path to be relative to root_dir
                        relative_path = os.path.relpath(directory_path, root_dir)
                        shortened_path = "./" if relative_path == "." else os.path.join("./", relative_path)

                        file_data.append([file_name, shortened_path, filepath, file_size, modified_time, None, "No"])  # Append to file_data with "No" for delete
                    except FileNotFoundError:
                        print(f"Warning: File not found: {filepath}")
                        directory_path, file_name = os.path.split(filepath)  # split the paths
                         # Shorten the directory path to be relative to root_dir
                        relative_path = os.path.relpath(directory_path, root_dir)
                        shortened_path = "./" if relative_path == "." else os.path.join("./", relative_path)

                        file_data.append([file_name, shortened_path, filepath, "N/A", "N/A", None,  "No"])
                    except OSError as e:
                        print(f"Warning: Error accessing file {filepath}: {e}")
                        directory_path, file_name = os.path.split(filepath)  # split the paths
                         # Shorten the directory path to be relative to root_dir
                        relative_path = os.path.relpath(directory_path, root_dir)
                        shortened_path = "./" if relative_path == "." else os.path.join("./", relative_path)

                        file_data.append([file_name, shortened_path, filepath, "N/A", "N/A", None, "No"])
                    except Exception as e:
                        print(f"Warning: An unexpected error occurred processing {filepath}: {e}")
                        directory_path, file_name = os.path.split(filepath)  # split the paths
                         # Shorten the directory path to be relative to root_dir
                        relative_path = os.path.relpath(directory_path, root_dir)
                        shortened_path = "./" if relative_path == "." else os.path.join("./", relative_path)

                        file_data.append([file_name, shortened_path, filepath, "N/A", "N/A", None, "No"])
                else:
                    directory_path, file_name = os.path.split(filepath) # split the paths
                    # Shorten the directory path to be relative to root_dir
                    relative_path = os.path.relpath(directory_path, root_dir)
                    shortened_path = "./" if relative_path == "." else os.path.join("./", relative_path)

                    file_data.append([file_name, shortened_path, filepath, "N/A", "N/A", None, "No"])  # Append to file_data (handles non-image or video)
                pbar.update(1)

    # Assign group numbers
    grouped_files = []
    unique_files = []
    group_counter = 1
    for hash_value, file_list in duplicate_groups.items():
        if len(file_list) > 1:  # Duplicate group
            for item in file_data:
                if item[2] in file_list:  # item[2] is the File Path
                    item[5] = group_counter # Group Number
                    grouped_files.append(item)
            group_counter += 1
        else:  # Unique file
            for item in file_data:
                if item[2] == file_list[0]:  # item[2] is the File Path
                    unique_files.append(item)

    # Write grouped files first, then unique files
    for row in grouped_files:
        sheet.append(row)
    # Add a blank row as a visual separator
    if grouped_files and unique_files:  # Only add separator if both groups exist
        sheet.append(["", "", "", "", "", "", ""]) # Empty row
    for row in unique_files:
        sheet.append(row)

    try:
        print(f"Saving to: {excel_file}")
        workbook.save(excel_file)
        print(f"File list saved to: {excel_file}")
    except PermissionError as e:
        print(f"Error: Unable to save file to {excel_file}. Check write permissions: {e}")
    except FileNotFoundError as e:
        print(f"Error: File or directory not found: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")


def is_image_or_video(filename):
    """Checks if a file is an image or video based on extension."""
    ext = filename.lower().split('.')[-1]
    image_extensions = ["jpg", "jpeg", "png", "gif", "bmp", "webp"]
    video_extensions = ["mp4", "avi", "mov", "mkv", "webm"]
    return ext in image_extensions or ext in video_extensions


def get_file_list_from_excel(excel_file):
    """Reads the Excel file and returns the file list as a list of dictionaries."""
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
    except FileNotFoundError:
        print(f"Error: Excel file not found: {excel_file}")
        return []
    except Exception as e:
        print(f"Error: Could not open Excel file: {e}")
        return []

    file_list = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Start from row 2 (skip header)
        if row:
            file_list.append({
                "File Name": row[0], #file name
                "Directory Path": row[1], #Directory Path
                "File Path": row[2], #original file path hidden
                "File Size (bytes)": row[3],
                "Last Modified": row[4],
                "Duplicate Group": row[5],
                "Delete": row[6] # Now only one column
            })
    return file_list


def save_file_list_to_excel(excel_file, file_list):
    """Saves the file list back to the Excel file."""
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Write header row
        header = list(file_list[0].keys()) if file_list else ["File Name", "Directory Path", "File Path", "File Size (bytes)", "Last Modified", "Duplicate Group", "Delete"] # Remove "Keep"
        sheet.append(header)

        # Write data rows
        for file_data in file_list:
            row = [file_data[key] for key in header]
            sheet.append(row)

        workbook.save(excel_file)
        print(f"File list saved to: {excel_file}")
    except Exception as e:
        print(f"Error saving file list to Excel: {e}")

def delete_marked_files(excel_file):
    """Deletes files based on the Excel file and "Delete" column."""
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
    except FileNotFoundError:
        print(f"Error: Excel file not found: {excel_file}")
        return jsonify({"status": "error", "message": "Excel file not found"})
    except Exception as e:
        print(f"Error: Could not open Excel file: {e}")
        return jsonify({"status": "error", "message": f"Could not open Excel file: {e}"})

    deleted_count = 0
    not_found_count = 0
    error_count = 0

    for row_index in tqdm(range(2, sheet.max_row + 1), desc="Deleting Files", disable=True): #disable tqdm on web
        file_path = sheet.cell(row=row_index, column=3).value  # Use the File Path column(colum 3)
        delete_flag = sheet.cell(row=row_index, column=7).value  # Check the Delete column change to 7

        if delete_flag and delete_flag.lower() == "yes":
            try:
                os.remove(file_path)
                deleted_count += 1
                print(f"Deleted: {file_path}")
            except FileNotFoundError:
                print(f"Warning: File not found (cannot delete): {file_path}")
                not_found_count += 1
            except Exception as e:
                print(f"Warning: Error deleting {file_path}: {e}")
                error_count += 1

    print(f"Deleted {deleted_count} files.")
    print(f"{not_found_count} files were not found.")
    print(f"{error_count} errors occurred during deletion.")

    return jsonify({"status": "success", "message": f"Deleted {deleted_count} files."})  # Return result

@app.route('/')
def index():
    """Serves the main page."""
    return render_template('index.html')  # Changed to render the template

@app.route('/set_root_directory', methods=['POST'])
def set_root_directory():
    global ROOT_DIRECTORY
    data = request.get_json()
    ROOT_DIRECTORY = data.get('rootDirectory', '')
    print(f"Root directory set to: {ROOT_DIRECTORY}")
    return jsonify({"status": "success", "message": "Root directory set successfully"})

@app.route('/list_files')
def list_files():
    """Lists files and returns data for the web interface."""
    global ROOT_DIRECTORY
    if not ROOT_DIRECTORY:
        return jsonify({"status": "error", "message": "Root directory is not set. Please set it first."})

    list_files_and_find_duplicates(ROOT_DIRECTORY, EXCEL_FILE)
    file_list = get_file_list_from_excel(EXCEL_FILE)
    return jsonify({"status": "success", "file_list": file_list})

@app.route('/update_file', methods=['POST'])
def update_file():
    """Updates the 'Delete' status of a file in the Excel file."""
    data = request.get_json()
    file_path = data.get('filePath')
    delete_value = data.get('delete')

    # Get the existing file list
    file_list = get_file_list_from_excel(EXCEL_FILE)

    # Find the file in the list
    for file_data in file_list:
        if file_data["File Path"] == file_path:
            # Update the Delete value
            file_data["Delete"] = delete_value
            break

    # Save the updated file list back to the Excel file
    save_file_list_to_excel(EXCEL_FILE, file_list)

    # Re-evaluate group statuses
    group_statuses = get_group_statuses(file_list)
    # Return data for refresh
    return jsonify({"status": "success", "message": f"Updated {file_path} with Delete={delete_value}",  "file_list":file_list,"group_statuses": group_statuses}) #update: new file with all set

def get_group_statuses(file_list):
    """Calculates and returns group statuses based on the file list."""
    duplicate_groups = {}
    for file_data in file_list:
        group_number = file_data.get('Duplicate Group')
        if group_number:
            if group_number not in duplicate_groups:
                duplicate_groups[group_number] = []
            duplicate_groups[group_number].append(file_data)

    group_statuses = {}
    for group_number, files in duplicate_groups.items():
        num_not_deleted = sum(1 for file_data in files if file_data["Delete"].lower() != "yes")

        if len(files) == 1:  # If only one file
            group_statuses[group_number] = "Only 1 file"  # Report as only 1 file
        elif num_not_deleted == 0:  # If there are multiple files, and ALL is on delete
            group_statuses[group_number] = "All to be Deleted"  # Mark group to delete everything
        elif num_not_deleted == 1:
            group_statuses[group_number] = "Only 1 file"
        else:
            group_statuses[group_number] = "Multiple Duplicates"  # Mark as having multiple duplicates

    return group_statuses

@app.route('/apply_folder_priority', methods=['POST'])
def apply_folder_priority():
    """Applies folder priority rules.  This version prioritizes folder and automatically sets Delete columns."""
    data = request.get_json()
    priority_folders = data.get('priorityFolders', [])

    # Get the file list from the Excel file
    file_list = get_file_list_from_excel(EXCEL_FILE)

    # Group files by duplicate group
    duplicate_groups = {}
    for file_data in file_list:
        group_number = file_data.get('Duplicate Group')
        if group_number:
            if group_number not in duplicate_groups:
                duplicate_groups[group_number] = []
            duplicate_groups[group_number].append(file_data)

    # Apply priority rules within each duplicate group
    group_statuses = {}  # Store status of each group
    for group_number, files in duplicate_groups.items():
        # Sort files based on folder priority
        sorted_files = sorted(files, key=lambda x: (
            priority_folders.index(x["Directory Path"]) if x["Directory Path"] in priority_folders else len(priority_folders), # Prioritize priority folder, or set lower
            x["File Path"]  # Secondary sort by filename (optional)
        ))

        # Set "Delete" to "Yes" for lower priority
        for i, file_data in enumerate(sorted_files):
            if i == 0:
                file_data["Delete"] = "No" # The first in priority is set to no delete
            else:
                file_data["Delete"] = "Yes"  # Mark lower-priority file for deletion

    # Save the updated file list back to the Excel file
    save_file_list_to_excel(EXCEL_FILE, file_list)

      # Re-evaluate group statuses and send back
    group_statuses = get_group_statuses(file_list)

    return jsonify({"status": "success", "message": "Folder priority applied successfully",  "file_list":file_list,"group_statuses": group_statuses})

@app.route('/delete_files')
def delete_files_route():
    """Deletes files marked for deletion."""
    result = delete_marked_files(EXCEL_FILE)  # Call delete_marked_files and store the result
    return result

@app.route('/get_image/<path:filename>')
def get_image(filename):
    """Serve image files from the specified directory."""
    # Ensure that the filename is safe (e.g., using os.path.basename)
    safe_filename = os.path.basename(filename)
    # Get the directory of the file from the 'root_directory' variable
    directory = os.path.dirname(filename)
    # Try to serve the file from the specified directory
    try:
        return send_from_directory(directory, safe_filename)
    except FileNotFoundError:
        # If the file is not found in the specified directory, return a 404 error
        abort(404)



if __name__ == '__main__':
    app.run(debug=True)